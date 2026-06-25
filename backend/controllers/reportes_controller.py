import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from models.producto import Producto
from models.venta import Venta

VERDE_IDEAL = colors.HexColor("#1A7A4C")
ROJO_IDEAL = colors.HexColor("#E2342B")


class ReportesController:

    @staticmethod
    def reporte_inventario():
        productos = Producto.listar(solo_activos=True)
        disponibles = [p for p in productos if p["stock"] > p["stock_minimo"]]
        bajo_stock = [p for p in productos if 0 < p["stock"] <= p["stock_minimo"]]
        agotados = [p for p in productos if p["stock"] <= 0]
        return {
            "disponibles": disponibles,
            "bajo_stock": bajo_stock,
            "agotados": agotados,
            "todos": productos,
        }

    @staticmethod
    def reporte_ventas(periodo="mensual"):
        return Venta.reporte_ganancias(periodo=periodo)

    @staticmethod
    def exportar_inventario_excel():
        productos = Producto.listar(solo_activos=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario"

        headers = ["ID", "Nombre", "Categoría", "Precio", "Stock", "Stock Mínimo", "Estado"]
        ws.append(headers)
        header_fill = PatternFill(start_color="1A7A4C", end_color="1A7A4C", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill

        for p in productos:
            estado = "Agotado" if p["stock"] <= 0 else ("Bajo" if p["stock"] <= p["stock_minimo"] else "Disponible")
            ws.append([
                p["id"], p["nombre"], p["categoria"], float(p["precio"]),
                p["stock"], p["stock_minimo"], estado,
            ])

        for col in ws.columns:
            max_len = max(len(str(c.value)) for c in col if c.value is not None)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def exportar_ventas_excel(periodo="mensual"):
        datos = Venta.reporte_ganancias(periodo=periodo)
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Ventas"

        headers = ["Periodo", "Número de Ventas", "Total Vendido"]
        ws.append(headers)
        header_fill = PatternFill(start_color="E2342B", end_color="E2342B", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill

        for d in datos:
            ws.append([d["periodo"], d["num_ventas"], float(d["total"] or 0)])

        for col in ws.columns:
            max_len = max(len(str(c.value)) for c in col if c.value is not None)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def exportar_inventario_pdf():
        data_reporte = ReportesController.reporte_inventario()
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=2 * cm, leftMargin=2 * cm,
                                 rightMargin=2 * cm, bottomMargin=2 * cm)
        styles = getSampleStyleSheet()
        elementos = [Paragraph("Reporte de Inventario — Salsa Ideal", styles["Title"]),
                     Spacer(1, 0.5 * cm)]

        data = [["Producto", "Categoría", "Stock", "Mínimo", "Estado"]]
        for p in data_reporte["todos"]:
            estado = "Agotado" if p["stock"] <= 0 else ("Bajo" if p["stock"] <= p["stock_minimo"] else "OK")
            data.append([p["nombre"], p["categoria"], str(p["stock"]), str(p["stock_minimo"]), estado])

        tabla = Table(data, colWidths=[7 * cm, 4 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm])
        tabla.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), VERDE_IDEAL),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F7F7")]),
        ]))
        elementos.append(tabla)
        doc.build(elementos)
        buffer.seek(0)
        return buffer

    @staticmethod
    def exportar_ventas_pdf(periodo="mensual"):
        datos = Venta.reporte_ganancias(periodo=periodo)
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=2 * cm, leftMargin=2 * cm,
                                 rightMargin=2 * cm, bottomMargin=2 * cm)
        styles = getSampleStyleSheet()
        elementos = [Paragraph(f"Reporte de Ventas ({periodo}) — Salsa Ideal", styles["Title"]),
                     Spacer(1, 0.5 * cm)]

        data = [["Periodo", "N° Ventas", "Total"]]
        for d in datos:
            data.append([d["periodo"], str(d["num_ventas"]), f"${(d['total'] or 0):,.2f}"])

        tabla = Table(data, colWidths=[6 * cm, 4 * cm, 6 * cm])
        tabla.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), ROJO_IDEAL),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F7F7")]),
        ]))
        elementos.append(tabla)
        doc.build(elementos)
        buffer.seek(0)
        return buffer
